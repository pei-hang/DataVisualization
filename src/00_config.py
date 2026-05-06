from __future__ import annotations

import math
import re
import warnings
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd


ROOT_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT_DIR / "data"
RAW_DIR = DATA_DIR / "raw"
INTERIM_DIR = DATA_DIR / "interim"
PROCESSED_DIR = DATA_DIR / "processed"
OUTPUTS_DIR = ROOT_DIR / "outputs"
TABLES_DIR = OUTPUTS_DIR / "tables"
FIGURES_DIR = OUTPUTS_DIR / "figures"
REPORTS_DIR = OUTPUTS_DIR / "reports"


def raw_file(*names: str) -> Path:
    for name in names:
        path = RAW_DIR / name
        if path.exists():
            return path
    return RAW_DIR / names[0]


BALANCE_FILE = raw_file("FS_Combas.xlsx", "BalanceTable.xlsx")
INCOME_FILE = raw_file("FS_Comins.xlsx", "IncomeTable.xlsx")
CASHFLOW_FILE = raw_file("FS_Comscfd.xlsx", "CashFlowTable.xlsx")
HEADCOUNT_FILE = raw_file("CG_Ybasic.xlsx", "HeadcountTable.xlsx")
INFO_FILE = raw_file("STK_LISTEDCOINFOANL.xlsx", "InfoTable.xlsx")
DFI_FILE = RAW_DIR / "北京大学数字普惠金融指数（PKU-DFIIC）2011-2023.xlsx"

PANEL_RAW_FILE = INTERIM_DIR / "merged_raw_panel.csv"
CITY_DFI_FILE = INTERIM_DIR / "dfi_city_clean.csv"
PROVINCE_DFI_FILE = INTERIM_DIR / "dfi_province_clean.csv"
PANEL_MODEL_FILE = PROCESSED_DIR / "panel_model_data.csv"
PANEL_TFP_FILE = PROCESSED_DIR / "panel_model_data_with_tfp.csv"

SAMPLE_START = 2011
SAMPLE_END = 2023
RANDOM_STATE = 20240502
FIG_DPI = 300

CSMAR_TABLES: dict[str, dict[str, Any]] = {
    "balance": {
        "path": BALANCE_FILE,
        "sheet": "sheet1",
        "max_col": None,
        "required": [
            "Stkcd",
            "ShortName",
            "Accper",
            "Typrep",
            "A001212000",
            "A001000000",
            "A002000000",
        ],
        "rename": {
            "Stkcd": "stkcd",
            "A001212000": "fixed_assets",
            "A001000000": "total_assets",
            "A002000000": "total_liabilities",
        },
        "date_col": "Accper",
        "keep_typrep_a": True,
    },
    "income": {
        "path": INCOME_FILE,
        "sheet": "sheet1",
        "max_col": None,
        "required": [
            "Stkcd",
            "ShortName",
            "Accper",
            "Typrep",
            "B001101000",
            "B001201000",
            "B001210000",
            "B001216000",
            "B002000000",
        ],
        "rename": {
            "Stkcd": "stkcd",
            "B001101000": "revenue",
            "B001201000": "operating_cost",
            "B001210000": "admin_expense",
            "B001216000": "rd_expense",
            "B002000000": "net_profit",
        },
        "date_col": "Accper",
        "keep_typrep_a": True,
    },
    "cashflow": {
        "path": CASHFLOW_FILE,
        "sheet": "sheet1",
        "max_col": None,
        "required": [
            "Stkcd",
            "ShortName",
            "Accper",
            "Typrep",
            "C001014000",
            "C001000000",
            "C002000000",
        ],
        "rename": {
            "Stkcd": "stkcd",
            "C001014000": "cash_paid_goods_services",
            "C001000000": "operating_cashflow",
            "C002000000": "investing_cashflow",
        },
        "date_col": "Accper",
        "keep_typrep_a": True,
    },
    "headcount": {
        "path": HEADCOUNT_FILE,
        "sheet": "sheet1",
        "max_col": None,
        "required": ["Stkcd", "Reptdt", "Y0601b"],
        "rename": {"Stkcd": "stkcd", "Y0601b": "employees"},
        "date_col": "Reptdt",
        "keep_typrep_a": False,
    },
    "info": {
        "path": INFO_FILE,
        "sheet": "sheet1",
        "max_col": None,
    },
}

DFI_CITY_COLUMNS = [
    "year",
    "pref_name_year18",
    "pref_name_year18_eng",
    "pref_code_year18",
    "pref_name_year14",
    "pref_code_year14",
    "index_aggregate",
    "coverage_breadth",
    "usage_depth",
    "payment",
    "insurance",
    "monetary_fund",
    "investment",
    "credit",
    "credit_investigation",
    "digitization_level",
]
DFI_PROVINCE_COLUMNS = [
    "year",
    "prov_name",
    "prov_code",
    "index_aggregate",
    "coverage_breadth",
    "usage_depth",
    "payment",
    "insurance",
    "monetary_fund",
    "investment",
    "credit",
    "credit_investigation",
    "digitization_level",
]
DFI_VALUE_COLUMNS = [
    "index_aggregate",
    "coverage_breadth",
    "usage_depth",
    "payment",
    "insurance",
    "monetary_fund",
    "investment",
    "credit",
    "credit_investigation",
    "digitization_level",
]

FINANCIAL_NUMERIC_COLUMNS = [
    "fixed_assets",
    "total_assets",
    "total_liabilities",
    "revenue",
    "operating_cost",
    "admin_expense",
    "rd_expense",
    "net_profit",
    "cash_paid_goods_services",
    "operating_cashflow",
    "investing_cashflow",
    "employees",
]

CONTROL_VARS = [
    "Size",
    "Lev",
    "ROA",
    "Growth",
    "Cashflow",
    "Tangibility",
    "RD",
    "rd_missing",
    "Age",
]
CONTROL_VARS_WITHOUT_RD = [
    "Size",
    "Lev",
    "ROA",
    "Growth",
    "Cashflow",
    "Tangibility",
    "rd_missing",
    "Age",
]
CONTROL_VARS_SA_MECHANISM = [
    "Lev",
    "ROA",
    "Growth",
    "Cashflow",
    "Tangibility",
    "RD",
    "rd_missing",
]
WINSOR_VARS = [
    "lnY",
    "lnK",
    "lnL",
    "lnM",
    "Size",
    "Lev",
    "ROA",
    "Growth",
    "Cashflow",
    "Tangibility",
    "RD",
    "RD_observed",
    "lnrd_expense",
    "AdminRatio",
    "InvestCashflow",
    "SA",
    "Turnover",
    "lndigital",
    "lndigital_coverage",
    "lndigital_depth",
    "lndigital_digitization",
]
ML_BASE_FEATURES = [
    "lndigital",
    "lndigital_coverage",
    "lndigital_depth",
    "lndigital_digitization",
    "Size",
    "Lev",
    "ROA",
    "Growth",
    "Cashflow",
    "Tangibility",
    "RD",
    "rd_missing",
    "SA",
    "Turnover",
    "AdminRatio",
    "Age",
    "SOE",
    "manufacturing",
    "hightech",
]

EAST_PROVINCES = {"北京", "天津", "河北", "上海", "江苏", "浙江", "福建", "山东", "广东", "海南"}
CENTRAL_PROVINCES = {"山西", "安徽", "江西", "河南", "湖北", "湖南"}
WEST_PROVINCES = {"内蒙古", "广西", "重庆", "四川", "贵州", "云南", "西藏", "陕西", "甘肃", "青海", "宁夏", "新疆"}
NORTHEAST_PROVINCES = {"辽宁", "吉林", "黑龙江"}
MUNICIPALITIES = {"北京", "上海", "天津", "重庆"}


def ensure_project_dirs() -> None:
    for path in [INTERIM_DIR, PROCESSED_DIR, TABLES_DIR, FIGURES_DIR, REPORTS_DIR]:
        path.mkdir(parents=True, exist_ok=True)


def require_columns(df: pd.DataFrame, required: list[str], dataset_name: str) -> None:
    missing = [col for col in required if col not in df.columns]
    if missing:
        raise ValueError(f"{dataset_name} 缺失字段: {missing}")


def normalize_stkcd(value: Any) -> str | pd.NA:
    if pd.isna(value):
        return pd.NA
    text = str(value).strip()
    if not text:
        return pd.NA
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    text = re.sub(r"\D", "", text)
    if not text:
        return pd.NA
    return text.zfill(6)


def infer_header_width(worksheet: Any, scan_limit: int = 300) -> int:
    first_row = [worksheet.cell(1, col).value for col in range(1, scan_limit + 1)]
    nonempty = [idx + 1 for idx, value in enumerate(first_row) if value not in (None, "")]
    if not nonempty:
        raise ValueError("第 1 行没有识别到英文字段名。")
    return max(nonempty)


def read_csmar_excel_fixed_cols(path: Path, sheet: str, max_col: int | None, dataset_name: str) -> pd.DataFrame:
    from openpyxl import load_workbook

    if not path.exists():
        raise FileNotFoundError(f"{dataset_name} 文件不存在: {path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"{dataset_name} 缺失工作表 {sheet}; 当前工作表: {workbook.sheetnames}")
    worksheet = workbook[sheet]
    if max_col is None:
        max_col = infer_header_width(worksheet)
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, max_col=max_col, values_only=True))
    headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
    if any(not header for header in headers):
        workbook.close()
        raise ValueError(f"{dataset_name} 第 1 行字段名存在空值，请检查前 {max_col} 列: {headers}")

    rows = list(worksheet.iter_rows(min_row=4, max_col=max_col, values_only=True))
    workbook.close()
    df = pd.DataFrame(rows, columns=headers)
    df = df.dropna(how="all").reset_index(drop=True)
    return df


def parse_datetime(series: pd.Series) -> pd.Series:
    return pd.to_datetime(series, errors="coerce")


def keep_year_end_sample(df: pd.DataFrame, date_col: str, dataset_name: str) -> pd.DataFrame:
    date_series = parse_datetime(df[date_col])
    out = df.copy()
    out[f"{date_col}_dt"] = date_series
    bad_dates = int(out[f"{date_col}_dt"].isna().sum())
    if bad_dates:
        warnings.warn(f"{dataset_name} 有 {bad_dates} 行日期无法解析，将在年度样本筛选中剔除。")
    out["year"] = out[f"{date_col}_dt"].dt.year
    out = out[out[f"{date_col}_dt"].dt.strftime("%m-%d") == "12-31"].copy()
    out = out[out["year"].between(SAMPLE_START, SAMPLE_END)].copy()
    out["year"] = out["year"].astype(int)
    return out


def normalize_region_name(value: Any) -> str | pd.NA:
    if pd.isna(value):
        return pd.NA
    text = str(value).strip()
    text = re.sub(r"\s+", "", text)
    text = text.replace("　", "")
    if not text or text.lower() in {"nan", "none"}:
        return pd.NA
    suffixes = [
        "维吾尔自治区",
        "壮族自治区",
        "回族自治区",
        "特别行政区",
        "自治区",
        "自治州",
        "地区",
        "省",
        "市",
        "盟",
    ]
    for suffix in suffixes:
        if text.endswith(suffix):
            text = text[: -len(suffix)]
            break
    aliases = {"北京市": "北京", "上海市": "上海", "天津市": "天津", "重庆市": "重庆"}
    return aliases.get(text, text)


def region_group(province_norm: Any) -> str | pd.NA:
    if pd.isna(province_norm):
        return pd.NA
    province = str(province_norm)
    if province in EAST_PROVINCES:
        return "east"
    if province in CENTRAL_PROVINCES:
        return "central"
    if province in WEST_PROVINCES:
        return "west"
    if province in NORTHEAST_PROVINCES:
        return "northeast"
    return pd.NA


def to_numeric(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    out = df.copy()
    for col in columns:
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors="coerce")
    return out


def safe_log(series: pd.Series) -> pd.Series:
    values = pd.to_numeric(series, errors="coerce")
    return np.log(values.where(values > 0))


def safe_divide(numerator: pd.Series, denominator: pd.Series) -> pd.Series:
    num = pd.to_numeric(numerator, errors="coerce")
    den = pd.to_numeric(denominator, errors="coerce")
    return num.divide(den.where(den != 0))


def winsorize_series(series: pd.Series, lower: float = 0.01, upper: float = 0.99) -> pd.Series:
    values = pd.to_numeric(series, errors="coerce")
    non_missing = values.dropna()
    if non_missing.empty:
        return values
    low = non_missing.quantile(lower)
    high = non_missing.quantile(upper)
    return values.clip(lower=low, upper=high)


def winsorize_dataframe(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    out = df.copy()
    for col in columns:
        if col not in out.columns:
            raise ValueError(f"缩尾变量缺失: {col}")
        out[col] = winsorize_series(out[col])
    return out


def significance_stars(pvalue: float | None) -> str:
    if pvalue is None or pd.isna(pvalue):
        return ""
    if pvalue < 0.01:
        return "***"
    if pvalue < 0.05:
        return "**"
    if pvalue < 0.10:
        return "*"
    return ""


def format_reg_cell(coef: float | None, se: float | None, pvalue: float | None) -> str:
    if coef is None or pd.isna(coef):
        return ""
    coef_text = f"{coef:.4f}{significance_stars(pvalue)}"
    se_text = "" if se is None or pd.isna(se) else f"({se:.4f})"
    return f"{coef_text}\n{se_text}" if se_text else coef_text


def _result_value(result: dict[str, Any], key: str, name: str) -> float | None:
    values = result.get(key, {})
    if name not in values:
        return None
    value = values[name]
    try:
        if pd.isna(value):
            return None
    except TypeError:
        pass
    return float(value)


def run_panel_fe(
    df: pd.DataFrame,
    dep_var: str,
    x_vars: list[str],
    model_name: str,
    cluster_var: str = "stkcd",
) -> dict[str, Any]:
    required = ["stkcd", "year", dep_var] + x_vars
    if cluster_var not in {"stkcd", "year"}:
        required.append(cluster_var)
    require_columns(df, required, model_name)

    work = df[required].copy()
    work["stkcd"] = work["stkcd"].astype(str)
    work["year"] = pd.to_numeric(work["year"], errors="coerce")
    for col in [dep_var] + x_vars:
        work[col] = pd.to_numeric(work[col], errors="coerce")
    work = work.dropna(subset=required).copy()
    work["year"] = work["year"].astype(int)
    if work.empty:
        raise ValueError(f"{model_name} 没有可用于回归的非缺失样本。")
    if work["stkcd"].nunique() < 2 or work["year"].nunique() < 2:
        raise ValueError(f"{model_name} 需要至少 2 个企业和 2 个年份。")

    try:
        from linearmodels.panel import PanelOLS
        from statsmodels.tools.tools import add_constant

        panel = work.set_index(["stkcd", "year"])
        y = panel[dep_var].astype(float)
        x = add_constant(panel[x_vars].astype(float), has_constant="add")
        model = PanelOLS(y, x, entity_effects=True, time_effects=True, drop_absorbed=True, check_rank=False)
        if cluster_var == "stkcd":
            result = model.fit(cov_type="clustered", cluster_entity=True)
        elif cluster_var == "year":
            result = model.fit(cov_type="clustered", cluster_time=True)
        else:
            cluster_codes = pd.Series(work[cluster_var].astype("category").cat.codes.to_numpy(), index=panel.index)
            clusters = pd.DataFrame({"cluster": cluster_codes}, index=panel.index)
            result = model.fit(cov_type="clustered", clusters=clusters)

        return {
            "model": model_name,
            "engine": "linearmodels.PanelOLS",
            "dep_var": dep_var,
            "x_vars": x_vars,
            "cluster_var": cluster_var,
            "nobs": int(result.nobs),
            "rsquared": float(getattr(result, "rsquared", np.nan)),
            "rsquared_within": float(getattr(result, "rsquared_within", np.nan)),
            "params": result.params.to_dict(),
            "std_errors": result.std_errors.to_dict(),
            "pvalues": result.pvalues.to_dict(),
            "notes": "",
        }
    except Exception as panel_error:
        from statsmodels.formula.api import ols

        formula = f"{dep_var} ~ {' + '.join(x_vars)} + C(stkcd) + C(year)"
        fitted = ols(formula, data=work).fit(cov_type="cluster", cov_kwds={"groups": work[cluster_var]})
        return {
            "model": model_name,
            "engine": "statsmodels.OLS_FE_fallback",
            "dep_var": dep_var,
            "x_vars": x_vars,
            "cluster_var": cluster_var,
            "nobs": int(fitted.nobs),
            "rsquared": float(fitted.rsquared),
            "rsquared_within": np.nan,
            "params": fitted.params.to_dict(),
            "std_errors": fitted.bse.to_dict(),
            "pvalues": fitted.pvalues.to_dict(),
            "notes": f"PanelOLS 失败后使用 statsmodels 备选。原错误: {panel_error}",
        }


def regression_results_to_frames(
    results: list[dict[str, Any]],
    ordered_vars: list[str],
) -> tuple[pd.DataFrame, pd.DataFrame]:
    ordered_vars = list(dict.fromkeys(ordered_vars))
    formatted = pd.DataFrame(
        index=ordered_vars + ["Observations", "R-squared", "Within R-squared", "Engine"],
        dtype=object,
    )
    long_rows: list[dict[str, Any]] = []

    for result in results:
        model_name = result["model"]
        for var in ordered_vars:
            coef = _result_value(result, "params", var)
            se = _result_value(result, "std_errors", var)
            pvalue = _result_value(result, "pvalues", var)
            formatted.loc[var, model_name] = format_reg_cell(coef, se, pvalue)
            long_rows.append(
                {
                    "model": model_name,
                    "dep_var": result.get("dep_var"),
                    "variable": var,
                    "coef": coef,
                    "std_error": se,
                    "p_value": pvalue,
                    "nobs": result.get("nobs"),
                    "rsquared": result.get("rsquared"),
                    "rsquared_within": result.get("rsquared_within"),
                    "cluster_var": result.get("cluster_var"),
                    "engine": result.get("engine"),
                    "notes": result.get("notes", ""),
                }
            )
        formatted.loc["Observations", model_name] = "" if result.get("nobs") is None else str(result.get("nobs"))
        formatted.loc["R-squared", model_name] = finite_or_blank(result.get("rsquared"))
        formatted.loc["Within R-squared", model_name] = finite_or_blank(result.get("rsquared_within"))
        formatted.loc["Engine", model_name] = str(result.get("engine", ""))

    long = pd.DataFrame(long_rows)
    return formatted, long


def save_regression_results(
    results: list[dict[str, Any]],
    ordered_vars: list[str],
    output_path: Path,
    summary_sheet: str = "formatted",
) -> tuple[pd.DataFrame, pd.DataFrame]:
    formatted, long = regression_results_to_frames(results, ordered_vars)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        formatted.to_excel(writer, sheet_name=summary_sheet)
        long.to_excel(writer, sheet_name="long", index=False)
    return formatted, long


def key_coefficient_frame(results: list[dict[str, Any]], variable: str = "lndigital") -> pd.DataFrame:
    rows = []
    for result in results:
        coef = _result_value(result, "params", variable)
        se = _result_value(result, "std_errors", variable)
        pvalue = _result_value(result, "pvalues", variable)
        if coef is None or se is None:
            continue
        rows.append(
            {
                "model": result["model"],
                "variable": variable,
                "coef": coef,
                "std_error": se,
                "p_value": pvalue,
                "ci_low": coef - 1.96 * se,
                "ci_high": coef + 1.96 * se,
                "nobs": result.get("nobs"),
                "engine": result.get("engine"),
            }
        )
    return pd.DataFrame(rows)


def write_markdown(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def append_markdown(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as file:
        file.write(content)


def finite_or_blank(value: Any) -> str:
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return ""
    if not math.isfinite(numeric):
        return ""
    return f"{numeric:.4f}"
